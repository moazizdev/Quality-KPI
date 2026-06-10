import io
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HEADER_FILL = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
HEADER_FONT = Font(color="ffffff", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def auto_width(ws, num_cols, min_width=10, max_width=40):
    for col in range(1, num_cols + 1):
        lengths = []
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value is not None:
                    lengths.append(len(str(cell.value)))
        best = max(lengths) + 2 if lengths else min_width
        ws.column_dimensions[chr(64 + col)].width = min(max(best, min_width), max_width)


def build_excel(filename: str, headers: list[str], rows: list[list]) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    style_header(ws, len(headers))
    auto_width(ws, len(headers))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
